"""
utils.py

Fonctions utilitaires du comparatif
Tournoi / HelloAsso
"""

import re
from openpyxl.utils import get_column_letter


# ==========================================================
# Recherche automatique des entêtes
# ==========================================================

def chercher_entetes(ws, nom_colonne):
    """
    Recherche automatiquement la ligne contenant
    une colonne donnée.

    Paramètres
    ----------
    ws : Worksheet openpyxl

    nom_colonne : str
        Exemple :
            "N° Licence"
            "Désignation"

    Retour
    ------
    tuple
        (numero_ligne, liste_des_entetes)
    """

    for row in ws.iter_rows():

        headers = []

        for cell in row:

            if cell.value is None:
                headers.append("")
            else:
                headers.append(str(cell.value).strip())

        if nom_colonne in headers:
            return row[0].row, headers

    raise Exception(
        f"Impossible de trouver la colonne '{nom_colonne}'."
    )


# ==========================================================
# Conversion des montants
# ==========================================================

def convertir_montant(valeur):
    """
    Convertit un montant Excel en float.

    Exemples acceptés :

        18
        18.5
        "18"
        "18 €"
        "18,50 €"

    Retourne toujours un float.
    """

    if valeur is None:
        return 0.0

    if isinstance(valeur, (int, float)):
        return float(valeur)

    texte = (
        str(valeur)
        .replace("€", "")
        .replace(" ", "")
        .replace(",", ".")
    )

    try:
        return float(texte)
    except ValueError:
        return 0.0


# ==========================================================
# Extraction numéro licence
# ==========================================================

def extraire_licence(designation):
    """
    Exemple :

        Dupont Samuel - Licence 5621 - T5,TS

    Retour :

        5621
    """

    m = re.search(
        r"Licence\s+(\d+)",
        str(designation)
    )

    if m:
        return m.group(1)

    return None


# ==========================================================
# Extraction du nom du joueur
# ==========================================================

def extraire_nom(designation):
    """
    Exemple :

        Dupont Samuel - Licence 5621 - T5

    Retour :

        Dupont Samuel
    """

    morceaux = str(designation).split("-")

    return morceaux[0].strip()


# ==========================================================
# Extraction des tableaux
# ==========================================================

def extraire_tableaux(texte):
    """
    Exemples :

        T3 (OK), T4 (OK), TS (OK)

    ou

        Dupont Samuel - Licence 5621 - T3,T4,TS

    Retour :

        {'T3','T4','TS'}
    """

    if texte is None:
        return set()

    tableaux = re.findall(
        r"\b(TS|TH|TF|T\d+)\b",
        str(texte).upper()
    )

    return set(tableaux)


# ==========================================================
# Ajustement automatique des colonnes
# ==========================================================

def ajuster_largeur_colonnes(ws, largeur_max=40):
    """
    Ajuste automatiquement la largeur
    des colonnes.
    """

    for colonne in ws.columns:

        longueur = 0

        lettre = get_column_letter(
            colonne[0].column
        )

        for cellule in colonne:

            if cellule.value is None:
                continue

            longueur = max(
                longueur,
                len(str(cellule.value))
            )

        ws.column_dimensions[lettre].width = min(
            longueur + 3,
            largeur_max
        )


# ==========================================================
# Conversion d'un ensemble de tableaux
# ==========================================================

def tableaux_vers_texte(tableaux):
    """
    Convertit

        {"TS","T4","T3"}

    en

        T3, T4, TS
    """

    if not tableaux:
        return ""

    return ", ".join(sorted(tableaux))


# ==========================================================
# Différence entre deux ensembles
# ==========================================================

def tableaux_manquants(tableaux_tournoi, tableaux_hello):
    """
    Tableaux présents dans le tournoi
    mais absents de HelloAsso.
    """

    return sorted(
        tableaux_tournoi - tableaux_hello
    )


def tableaux_en_trop(tableaux_tournoi, tableaux_hello):
    """
    Tableaux présents dans HelloAsso
    mais absents du tournoi.
    """

    return sorted(
        tableaux_hello - tableaux_tournoi
    )


# ==========================================================
# Format Euro
# ==========================================================

def euro(valeur):
    """
    Formate un montant.

    Exemple :

        18

    devient

        18.00 €
    """

    return f"{valeur:.2f} €"


