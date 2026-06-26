"""
excel.py
Création du fichier Excel de comparaison.
"""

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from utils import ajuster_largeur_colonnes

VERT = PatternFill(fill_type="solid", fgColor="34EB57")
ORANGE = PatternFill(fill_type="solid", fgColor="CFAD1F")
ROUGE = PatternFill(fill_type="solid", fgColor="EB132E")
BLEU = PatternFill(fill_type="solid", fgColor="5207F7")
ENTETE = PatternFill(fill_type="solid", fgColor="0E3C8F")


def creer_excel(comparatif, stats, fichier):

    wb = Workbook()
    ws = wb.active
    ws.title = "Comparatif"

    entetes = [
        "Licence",
        "Joueur",
        "Tableaux Tournoi",
        "Tableaux HelloAsso",
        "Manquants",
        "En trop",
        "Montant Tournoi",
        "Montant HelloAsso",
        "Écart",
        "Statut"
    ]

    ws.append(entetes)

    # Entête
    for cell in ws[1]:
        cell.fill = ENTETE
        cell.font = Font(bold=True, color="FFFFFF")

    # Données
    for ligne in comparatif:

        ws.append([
            ligne["licence"],
            ligne["joueur"],
            ligne["tableaux_tournoi"],
            ligne["tableaux_hello"],
            ligne["manquants"],
            ligne["en_trop"],
            ligne["montant_tournoi"],
            ligne["montant_hello"],
            ligne["ecart"],
            ligne["statut"]
        ])

        r = ws.max_row

        statut = ligne["statut"]

        if statut == "OK":
            couleur = VERT
        elif statut == "Pas de paiement":
            couleur = ROUGE
        elif statut == "Paiement sans inscription":
            couleur = BLEU
        else:
            couleur = ORANGE

        for c in range(1, 11):
            ws.cell(r, c).fill = couleur

        # format €
        for c in (7, 8, 9):
            ws.cell(r, c).number_format = '#,##0.00 €'

    # Mise en forme
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    ajuster_largeur_colonnes(ws)

    # Statistiques
    l = ws.max_row + 3

    ws.cell(l, 1).value = "STATISTIQUES"
    ws.cell(l, 1).font = Font(bold=True)

    resume = [
        ("Comparaisons OK", stats["ok"]),
        ("Différences", stats["difference"]),
        ("Sans paiement", stats["sans_paiement"]),
        ("Paiements orphelins", stats["orphelin"]),
        ("", ""),
        ("Total Tournoi", stats["total_tournoi"]),
        ("Total HelloAsso", stats["total_hello"]),
        ("Écart Global", stats["total_hello"] - stats["total_tournoi"]),
    ]

    for titre, valeur in resume:
        l += 1
        ws.cell(l, 1).value = titre
        ws.cell(l, 2).value = valeur
        if isinstance(valeur, (int, float)) and "Total" in titre or "Écart" in titre:
            ws.cell(l, 2).number_format = '#,##0.00 €'

    wb.save(fichier)
