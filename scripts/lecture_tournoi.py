"""
lecture_tournoi.py

Lecture du fichier Inscriptions Tournoi.
"""

from openpyxl import load_workbook

from config import (
    FEUILLE_TOURNOI,
    COLONNE_LICENCE,
    COLONNE_JOUEUR,
    COLONNE_TABLEAUX,
    COLONNE_PAIEMENT,
    COLONNE_MONTANT,
)

from utils import (
    chercher_entetes,
    convertir_montant,
    extraire_tableaux,
)


# ==========================================================
# Lecture du fichier tournoi
# ==========================================================

def lire_tournoi(fichier):
    """
    Lit le fichier Inscriptions Tournoi.

    Retourne un dictionnaire :

    tournoi = {
        "5621": {
            "joueur": "...",
            "tableaux": {"T5","TS"},
            "tableaux_txt": "...",
            "paiement": "...",
            "montant": 28.0
        }
    }
    """

    print("\nLecture du fichier Tournoi...")

    wb = load_workbook(
        fichier,
        data_only=True
    )

    # --------------------------------------------------
    # Vérification feuille
    # --------------------------------------------------

    if FEUILLE_TOURNOI not in wb.sheetnames:

        raise Exception(
            f"La feuille '{FEUILLE_TOURNOI}' est introuvable."
        )

    ws = wb[FEUILLE_TOURNOI]

    # --------------------------------------------------
    # Recherche automatique des entêtes
    # --------------------------------------------------

    header_row, headers = chercher_entetes(
        ws,
        COLONNE_LICENCE
    )

    print(f"Entêtes trouvés ligne {header_row}")

    # --------------------------------------------------
    # Recherche des colonnes
    # --------------------------------------------------

    idx_licence = headers.index(COLONNE_LICENCE)
    idx_joueur = headers.index(COLONNE_JOUEUR)
    idx_tableaux = headers.index(COLONNE_TABLEAUX)
    idx_paiement = headers.index(COLONNE_PAIEMENT)
    idx_montant = headers.index(COLONNE_MONTANT)

    tournoi = {}

    # --------------------------------------------------
    # Lecture des joueurs
    # --------------------------------------------------

    for row in ws.iter_rows(
        min_row=header_row + 1,
        values_only=True
    ):

        licence = row[idx_licence]

        # Fin du tableau

        if licence is None:
            break

        licence = str(licence).strip()

        joueur = row[idx_joueur]

        tableaux_txt = row[idx_tableaux] or ""

        paiement = row[idx_paiement] or ""

        montant = convertir_montant(
            row[idx_montant]
        )

        tournoi[licence] = {

            "licence": licence,

            "joueur": joueur,

            "tableaux_txt": tableaux_txt,

            "tableaux": extraire_tableaux(
                tableaux_txt
            ),

            "paiement": paiement,

            "montant": montant

        }

    print(
        f"{len(tournoi)} joueurs chargés."
    )

    return tournoi

