"""
Lecture du fichier Export HelloAsso.

"""

from openpyxl import load_workbook

from config import (
    FEUILLE_HELLOASSO,
    COLONNE_DESIGNATION,
    COLONNE_TOTAL,
    COLONNE_STATUT,
    STATUT_VALIDE
)

from utils import (
    chercher_entetes,
    convertir_montant,
    extraire_licence,
    extraire_nom,
    extraire_tableaux
)

# Lecture du fichier HelloAsso

def lire_helloasso(fichier):
    
    print("\nLecture du fichier HelloAsso...")
    wb = load_workbook(
        fichier,
        data_only=True
    )

    # Vérification de la feuille

    if FEUILLE_HELLOASSO not in wb.sheetnames:

        raise Exception(
            f"La feuille '{FEUILLE_HELLOASSO}' est introuvable."
        )

    ws = wb[FEUILLE_HELLOASSO]

    # Recherche des entêtes

    header_row, headers = chercher_entetes(
        ws,
        COLONNE_DESIGNATION
    )
    print(f"Entêtes trouvés ligne {header_row}")

    # Recherche des colonnes

    idx_designation = headers.index(COLONNE_DESIGNATION)
    idx_total = headers.index(COLONNE_TOTAL)
    idx_statut = headers.index(COLONNE_STATUT)

    hello = {}

    # Lecture des paiements

    for row in ws.iter_rows(
        min_row=header_row + 1,
        values_only=True
    ):

        designation = row[idx_designation]
        if designation is None:
            continue
        # uniquement les paiements validés
        statut = str(row[idx_statut]).strip()
        if statut != STATUT_VALIDE:
            continue
        licence = extraire_licence(
            designation
        )
        if licence is None:
            continue
        joueur = extraire_nom(
            designation
        )
        tableaux = extraire_tableaux(
            designation
        )
        montant = convertir_montant(
            row[idx_total]
        )

        # Plusieurs paiements pour une même licence

        if licence not in hello:
            hello[licence] = {
                "licence": licence,
                "joueur": joueur,
                "tableaux": set(),
                "tableaux_txt": "",
                "montant": 0.0
            }

        hello[licence]["montant"] += montant
        hello[licence]["tableaux"].update(
            tableaux
        )


    # Construction du texte des tableaux

    for licence in hello:

        hello[licence]["tableaux_txt"] = ", ".join(
            sorted(
                hello[licence]["tableaux"]
            )
        )
    print(
        f"{len(hello)} paiements chargés."
    )
    return hello

