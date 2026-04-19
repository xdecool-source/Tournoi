# Export Excel 
# Preparation Feuille

from collections import defaultdict
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from core.config import TABLEAUX

#  Styles  

header_fill = PatternFill(start_color="FFFF00", fill_type="solid")
header_font = Font(bold=True)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
def auto_adjust_width(ws):
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2

def format_sheet(ws):
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(
                horizontal=cell.alignment.horizontal,
                vertical="center"
            )
    auto_adjust_width(ws)

def build_data(rows):
    data_by_table = defaultdict(list)
    data_joueurs = defaultdict(lambda: {
        "Licence": "",
        "Nom": "",
        "Classement": "",
        "Club": "",
        "Mail": "",
        "Inscriptions": []
    })

    for row in rows:
        r = dict(row)

        data_by_table[r["tableau"]].append({
            "Dossard": r["dossard"],
            "Licence": r["Licence"],
            "Nom": r["Nom Prénom"],
            "Classement": r["Classement"],
            "Club": r["Club"],
            "Mail": r["Mail"],
            "Statut": r["statut"]
        })

        joueur = data_joueurs[r["dossard"]]
        joueur["Licence"] = r["Licence"]
        joueur["Nom"] = r["Nom Prénom"]
        joueur["Classement"] = r["Classement"]
        joueur["Club"] = r["Club"]
        joueur["Mail"] = r["Mail"]
        joueur["Inscriptions"].append((r["tableau"], r["statut"]))

    return data_by_table, data_joueurs

#  Creation des joueurs dans la feuille Joueurs

def create_players_sheet(wb, data_joueurs):
    
    ws = wb.create_sheet("Joueurs")
    ws["A1"] = f"Total joueurs : {len(data_joueurs)}"
    ws["A1"].font = Font(bold=True)
    ws.append([])
    headers = ["Dossard","Licence","Nom Prénom","Classement","Club","Mail","Tableaux"]
    ws.append(headers)
    for col in ws[3]:
        col.fill = header_fill
        col.font = header_font

    for dossard, infos in sorted(data_joueurs.items()):
        ws.append([
            dossard,
            infos["Licence"],
            infos["Nom"],
            infos["Classement"],
            infos["Club"],
            infos["Mail"],
            ", ".join([f"{t} ({s})" for t, s in infos["Inscriptions"]])
        ])
    format_sheet(ws)

#  Création de la feuille Tableaux 

def create_table_sheets(wb, data_by_table):
    for tableau, joueurs in sorted(data_by_table.items()):
        ws = wb.create_sheet(tableau)

        joueurs_sorted = sorted(joueurs, key=lambda x: x["Dossard"])
        ws["A1"] = f"Total joueurs : {len(joueurs_sorted)}"
        ws.append([])
        headers = ["Dossard","Licence","Nom Prénom","Classement","Club","Mail","Statut"]
        ws.append(headers)
        for col in ws[3]:
            col.fill = header_fill
            col.font = header_font
        for joueur in joueurs_sorted:
            ws.append([
                joueur["Dossard"],
                joueur["Licence"],
                joueur["Nom"],
                joueur["Classement"],
                joueur["Club"],
                joueur["Mail"],
                joueur["Statut"]
            ])
        format_sheet(ws)
        ws["A3"].alignment = Alignment(horizontal="center")
        for row in range(4, ws.max_row + 1):
            ws[f"A{row}"].alignment = Alignment(horizontal="center")

#  Création des feuilles Tableaux 

def create_tableaux_sheet(wb, data_by_table):
    ws = wb.create_sheet("Tableaux")
    headers = [
        "Tableau",
        "Points min",
        "Points max",
        "Capacité",
        "Liste attente max",
        "Prix (€)",
        "Nb inscrits",
        "Nb validés",
        "Nb attente"
    ]

#  En-têtes
    for col_index, value in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_index, value=value)
        cell.font = Font(bold=True)
    row = 2
    light_gray = PatternFill(start_color="F2F2F2", fill_type="solid")

#  On parcourt tous les tableaux définis dans config
    for index, (tableau, config) in enumerate(sorted(TABLEAUX.items())):
        joueurs = data_by_table.get(tableau, [])
        prix = TABLEAUX.get(tableau, {}).get("prix", 0)
        nb_inscrits = len(joueurs)
        nb_valides = sum(1 for j in joueurs if j["Statut"].upper() == "OK")
        nb_attente = sum(1 for j in joueurs if j["Statut"].upper() == "ATTENTE")
        ws.cell(row=row, column=1, value=tableau)
        ws.cell(row=row, column=2, value=config["min"])
        ws.cell(row=row, column=3, value=config["max"])
        ws.cell(row=row, column=4, value=config["capacite"])
        ws.cell(row=row, column=5, value=config["attente"])
        ws.cell(row=row, column=6, value=prix)
        ws.cell(row=row, column=7, value=nb_inscrits)
        ws.cell(row=row, column=8, value=nb_valides)
        ws.cell(row=row, column=9, value=nb_attente)
        ws.cell(row=row, column=6).number_format = '#,##0.00 €'

# Ligne alternée gris clair
        if index % 2 == 0:
            for col in range(1, 10):
                ws.cell(row=row, column=col).fill = light_gray
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
        row += 1

#  Ajustement largeur colonnes
    
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2
        
        