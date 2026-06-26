"""
1. Calcule les montants
les inscriptions validées (OK) ;
les inscriptions en attente (ATTENTE).
Les prix sont lus dans config.py
TABLEAUX

2. Calcule les totaux
le montant total encaissé ;
le montant des inscriptions en attente.

3. Crée la feuille Prix
génère un tableau contenant :
le dossard ;
le nom ;
la licence ;
les tableaux ;
le mode de paiement ;
le montant validé ;
le montant en attente.
ajoute également le logo du tournoi.

4. Met en forme
applique :
les formats monétaires (€) ;
les lignes alternées ;
les colonnes ajustées automatiquement ;
le gel de l'en-tête (freeze_panes).

"""

from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.drawing.image import Image
from core.config import TABLEAUX
from pathlib import Path


def sanitize_excel(value):
    if isinstance(value, str):
        value = value.strip()
        if value.startswith(("=", "+", "-", "@")):
            return "'" + value
    return value


def create_price_sheet(wb, data_joueurs, root_dir):

    ws = wb.create_sheet("Prix")

    # Logo Tournoi

    image_path = root_dir / "ressources" / "Tournoi.jpg"
    if image_path.exists():
        img = Image(str(image_path))
        img.width = min(img.width, 600)
        img.height = 110
        ws.add_image(img, "C1")

    start_row = 8
    total_valide_general = 0
    total_attente_general = 0
    lignes_prix = []

    # Calcul des montants

    for dossard, infos in sorted(data_joueurs.items()):

        montant_valide = 0
        montant_attente = 0

        for tableau, statut in infos["Inscriptions"]:

            prix = TABLEAUX.get(tableau, {}).get("prix", 0)
            statut = (statut or "").upper()

            if statut == "OK":
                montant_valide += prix

            elif statut == "ATTENTE":
                montant_attente += prix

        total_valide_general += montant_valide
        total_attente_general += montant_attente
        
        """
        print(
            infos.get("Nom"),
            infos.get("paiement"),
            infos.keys()
            )
        """
        
        lignes_prix.append([
            dossard,
            sanitize_excel(infos["Nom"]),
            sanitize_excel(infos.get("Licence", "")),
            sanitize_excel(
                ", ".join(
                    [f"{t} ({s})" for t, s in infos["Inscriptions"]]
                )
            ),
            sanitize_excel(infos.get("paiement", "")),
            montant_valide,
            montant_attente
        ])

    # Totaux

    ws[f"A{start_row}"] = "Montants par joueur"
    ws[f"A{start_row}"].font = Font(bold=True, size=14)

    ws[f"A{start_row+1}"] = "Total encaissé (validés) :"
    ws[f"A{start_row+2}"] = "Total en attente :"

    ws[f"F{start_row+1}"] = total_valide_general
    ws[f"G{start_row+2}"] = total_attente_general

    ws[f"F{start_row+1}"].number_format = '#,##0.00 €'
    ws[f"G{start_row+2}"].number_format = '#,##0.00 €'

    ws[f"F{start_row+1}"].font = Font(bold=True)
    ws[f"G{start_row+2}"].font = Font(bold=True)

    ws[f"A{start_row+1}"].font = Font(bold=True)
    ws[f"A{start_row+2}"].font = Font(bold=True)

    # Tableau détaillé

    header_row = start_row + 4

    headers = [
        "Dossard",
        "Nom Prénom",
        "N° Licence",
        "Tableaux",
        "Paiement",
        "Montant validé (€)",
        "Montant en attente (€)"
    ]

    for col_index, value in enumerate(headers, 1):
        ws.cell(
            row=header_row,
            column=col_index,
            value=value
        ).font = Font(bold=True)

    data_row = header_row + 1

    light_gray = PatternFill(
        start_color="F2F2F2",
        fill_type="solid"
    )

    for index, ligne in enumerate(lignes_prix):

        for col_index, value in enumerate(ligne, 1):
            ws.cell(
                row=data_row,
                column=col_index,
                value=value
            )

        ws[f"B{data_row}"].font = Font(bold=True)

        ws[f"F{data_row}"].number_format = '#,##0.00 €'
        ws[f"G{data_row}"].number_format = '#,##0.00 €'

        ws[f"F{data_row}"].font = Font(bold=True)

        if index % 2 == 0:
            for col in range(1, 8):
                ws.cell(
                    row=data_row,
                    column=col
                ).fill = light_gray

        ws[f"A{data_row}"].alignment = Alignment(
            horizontal="center"
        )

        data_row += 1

    ws.freeze_panes = f"A{header_row+1}"

    # Auto largeur

    for col in ws.columns:

        max_length = 0
        col_letter = col[0].column_letter

        for cell in col:
            if cell.value:
                max_length = max(
                    max_length,
                    len(str(cell.value))
                )

        ws.column_dimensions[col_letter].width = max_length + 2